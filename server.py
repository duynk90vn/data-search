import json
import mimetypes
import os
import re
import sqlite3
import subprocess
import sys
import threading
import time
import unicodedata
import urllib.parse
import hashlib
import webbrowser
from datetime import datetime
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
STATIC_DIR = ROOT / "static"
CONFIG_DIR = ROOT / "config"
DATA_DIR = ROOT / "data"
DB_PATH = DATA_DIR / "bom_index.sqlite3"
SETTINGS_PATH = CONFIG_DIR / "settings.json"
TERMINOLOGY_PATH = CONFIG_DIR / "terminology.json"
DEFAULT_BOM_DIR = r"D:\10.Project\BOM"
DEFAULT_ADMIN_USER = "duynk90"
DEFAULT_ADMIN_PASSWORD = "08011994"
HIDDEN_MODEL_NAMES = {"tonghopbom", "tonghopbomcapnhat"}

index_lock = threading.Lock()
index_state = {"running": False, "message": "Idle", "indexed": 0, "total": 0, "errors": []}


def ensure_dirs():
    CONFIG_DIR.mkdir(exist_ok=True)
    DATA_DIR.mkdir(exist_ok=True)
    if not SETTINGS_PATH.exists():
        SETTINGS_PATH.write_text(
            json.dumps(
                {
                    "bom_folder": DEFAULT_BOM_DIR,
                    "admin_user": DEFAULT_ADMIN_USER,
                    "admin_password_hash": hashlib.sha256(DEFAULT_ADMIN_PASSWORD.encode("utf-8")).hexdigest(),
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
    if not TERMINOLOGY_PATH.exists():
        TERMINOLOGY_PATH.write_text(json.dumps(default_terms(), ensure_ascii=False, indent=2), encoding="utf-8")


def default_terms():
    return {
        "giá treo": ["吊架", "三角架"],
        "vít giá treo": ["吊架螺丝", "吊架螺釘", "吊架螺钉", "三角架螺絲", "三角架螺丝"],
        "tụ điện": ["電容", "电容"],
        "motor": ["馬達", "马达", "電機", "电机"],
        "mô tơ": ["馬達", "马达", "電機", "电机"],
        "công tắc": ["開關", "开关"],
        "chụp treo": ["吊鐘", "吊钟"],
        "cánh quạt": ["扇葉", "扇叶"],
        "miếng đệm cao su motor": ["馬達橡膠墊片", "马达橡胶垫片", "馬達膠墊", "马达胶垫"],
        "đệm cao su": ["橡膠墊片", "橡胶垫片", "膠墊", "胶垫"],
        "dây điện": ["電線", "电线", "線材", "线材"],
        "ốc vít": ["螺絲", "螺丝", "螺釘", "螺钉"],
        "tem cảnh báo": ["警告標", "警示標", "警语标"],
        "túi pe": ["PE袋"],
    }


def load_json(path, fallback):
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return fallback


def save_json(path, data):
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def normalize_text(value):
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text).lower()
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[\s\-_./\\()（）\[\]{}]+", "", text)


def normalized_words(value):
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text).lower()
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^\w\u4e00-\u9fff]+", " ", text, flags=re.UNICODE)
    return [word for word in text.split() if word]


def phrase_contains(container, phrase):
    container_words = normalized_words(container)
    phrase_words = normalized_words(phrase)
    if not phrase_words or len(phrase_words) == 1:
        return False
    for i in range(0, len(container_words) - len(phrase_words) + 1):
        if container_words[i : i + len(phrase_words)] == phrase_words:
            return True
    return False


def display_value(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def init_db():
    ensure_dirs()
    with db() as conn:
        conn.executescript(
            """
            PRAGMA journal_mode=WAL;
            CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT NOT NULL);
            CREATE TABLE IF NOT EXISTS files (
                id INTEGER PRIMARY KEY,
                path TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                model TEXT NOT NULL,
                size INTEGER NOT NULL,
                mtime REAL NOT NULL,
                row_count INTEGER DEFAULT 0,
                indexed_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS rows (
                id INTEGER PRIMARY KEY,
                file_id INTEGER NOT NULL,
                model TEXT NOT NULL,
                file_name TEXT NOT NULL,
                file_path TEXT NOT NULL,
                sheet TEXT NOT NULL,
                row_number INTEGER NOT NULL,
                bom_level INTEGER,
                part_no TEXT,
                name_cn TEXT,
                name_vi TEXT,
                pinyin TEXT,
                quantity TEXT,
                specification TEXT,
                headers_json TEXT NOT NULL,
                raw_row_json TEXT NOT NULL,
                search_text TEXT NOT NULL,
                normalized_text TEXT NOT NULL,
                FOREIGN KEY(file_id) REFERENCES files(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS idx_rows_file ON rows(file_id);
            CREATE INDEX IF NOT EXISTS idx_rows_model ON rows(model);
            CREATE INDEX IF NOT EXISTS idx_rows_norm ON rows(normalized_text);
            """
        )
        existing_cols = {r["name"] for r in conn.execute("PRAGMA table_info(rows)")}
        migrations = {
            "bom_level": "ALTER TABLE rows ADD COLUMN bom_level INTEGER",
            "quantity": "ALTER TABLE rows ADD COLUMN quantity TEXT",
            "specification": "ALTER TABLE rows ADD COLUMN specification TEXT",
        }
        for col, sql in migrations.items():
            if col not in existing_cols:
                conn.execute(sql)


def get_settings():
    settings = load_json(SETTINGS_PATH, {"bom_folder": DEFAULT_BOM_DIR})
    settings.setdefault("bom_folder", DEFAULT_BOM_DIR)
    settings.setdefault("admin_user", DEFAULT_ADMIN_USER)
    settings.setdefault("admin_password_hash", hashlib.sha256(DEFAULT_ADMIN_PASSWORD.encode("utf-8")).hexdigest())
    save_json(SETTINGS_PATH, settings)
    return settings


def public_settings():
    settings = get_settings()
    return {"bom_folder": settings["bom_folder"], "admin_user": settings["admin_user"]}


def check_auth(payload):
    settings = get_settings()
    username = str(payload.get("username") or "")
    password = str(payload.get("password") or "")
    password_hash = hashlib.sha256(password.encode("utf-8")).hexdigest()
    return username == settings.get("admin_user") and password_hash == settings.get("admin_password_hash")


def model_from_filename(path):
    stem = path.stem
    match = re.search(r"\(([^()]+)\)$", stem)
    code = match.group(1) if match else ""
    before_cn = re.split(r"[\u4e00-\u9fff]", stem, maxsplit=1)[0].strip("-_ ")
    if before_cn:
        return before_cn
    return code or stem


def header_map(ws):
    best_row = 1
    best_score = -1
    for r in range(1, min(ws.max_row, 15) + 1):
        values = [display_value(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        score = sum(1 for v in values if v in {"料件編號", "料件编号", "品名", "用量", "規格", "规格", "供應商名稱", "供应商名称"})
        if score > best_score:
            best_score = score
            best_row = r
    headers = {}
    last = ""
    for c in range(1, ws.max_column + 1):
        value = display_value(ws.cell(best_row, c).value).strip()
        if value:
            last = value
        label = value or f"{last}_{c}" if last else f"Column_{c}"
        headers[c] = label
    name_col = next((c for c, v in headers.items() if "品名" in v), None)
    qty_col = next((c for c, v in headers.items() if "用量" in v), None)
    spec_col = next((c for c, v in headers.items() if "規格" in v or "规格" in v), None)
    return best_row, headers, name_col, qty_col, spec_col


def merged_lookup(ws):
    lookup = {}
    for area in ws.merged_cells.ranges:
        value = ws.cell(area.min_row, area.min_col).value
        for row in range(area.min_row, area.max_row + 1):
            for col in range(area.min_col, area.max_col + 1):
                lookup[(row, col)] = value
    return lookup


def cell_value(ws, lookup, row, col):
    value = ws.cell(row, col).value
    if value is None and (row, col) in lookup:
        value = lookup[(row, col)]
    return display_value(value).strip()


def direct_cell_value(ws, row, col):
    return display_value(ws.cell(row, col).value).strip()


def hierarchy_columns(ws, header_row, name_col):
    cols = []
    if not name_col:
        return cols
    for c in range(1, name_col):
        header = direct_cell_value(ws, header_row, c)
        if "料件" in header or not header:
            has_code = False
            for r in range(header_row + 1, min(ws.max_row, header_row + 80) + 1):
                value = direct_cell_value(ws, r, c)
                if value and re.search(r"\d", value):
                    has_code = True
                    break
            if has_code:
                cols.append(c)
    return cols


def row_level_and_part_no(ws, row_idx, hierarchy_cols):
    for idx, col in enumerate(hierarchy_cols, start=1):
        value = direct_cell_value(ws, row_idx, col)
        if value and re.search(r"\d", value):
            return min(idx, 3), value
    return None, ""


def row_record(ws, lookup, row_idx, headers, name_col, qty_col, spec_col):
    raw = {}
    cells = []
    for c in range(1, ws.max_column + 1):
        value = cell_value(ws, lookup, row_idx, c)
        key = headers.get(c, f"Column_{c}")
        if value:
            raw[key] = value
            raw[f"__C{c}"] = value
            cells.append(value)
    if not cells:
        return None
    name_cn = cell_value(ws, lookup, row_idx, name_col) if name_col else ""
    pinyin = ""
    name_vi = ""
    if qty_col and qty_col > (name_col or 0):
        maybe_pinyin = cell_value(ws, lookup, row_idx, qty_col + 1)
        maybe_vi = cell_value(ws, lookup, row_idx, qty_col + 2)
        if maybe_pinyin and not re.search(r"[\u4e00-\u9fff]", maybe_pinyin) and re.search(r"[A-Za-zāáǎàēéěèīíǐìōóǒòūúǔùǖǘǚǜü]", maybe_pinyin):
            pinyin = maybe_pinyin
        if maybe_vi and not re.search(r"[\u4e00-\u9fff]", maybe_vi):
            name_vi = maybe_vi
    if name_col and (not pinyin or not name_vi):
        for c in range(name_col + 1, min(ws.max_column, name_col + 7) + 1):
            value = cell_value(ws, lookup, row_idx, c)
            if not value:
                continue
            lower = value.lower()
            if not name_vi and (re.search(r"[à-ỹăâêôơưđ]", lower) or any(w in lower for w in ["linh kiện", "quạt", "chụp", "giá", "vít", "cụm"])):
                name_vi = value
            elif not pinyin and not value.isdigit() and not re.search(r"[\u4e00-\u9fff]", value) and re.fullmatch(r"[A-Za-z0-9\s'\-āáǎàēéěèīíǐìōóǒòūúǔùǖǘǚǜü:]+", value):
                pinyin = value
    quantity = cell_value(ws, lookup, row_idx, qty_col) if qty_col else ""
    specification = cell_value(ws, lookup, row_idx, spec_col) if spec_col else ""
    search_text = " ".join(cells)
    return {
        "part_no": "",
        "bom_level": None,
        "name_cn": name_cn,
        "name_vi": name_vi,
        "pinyin": pinyin,
        "quantity": quantity,
        "specification": specification,
        "headers": [headers.get(c, f"Column_{c}") for c in range(1, ws.max_column + 1)],
        "raw": raw,
        "search_text": search_text,
        "normalized_text": normalize_text(search_text),
    }


def parse_workbook(path):
    model = model_from_filename(path)
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=True)
    parsed = []
    for ws in workbook.worksheets:
        if ws.max_row < 2 or ws.max_column < 2:
            continue
        lookup = merged_lookup(ws)
        header_row, headers, name_col, qty_col, spec_col = header_map(ws)
        hierarchy_cols = hierarchy_columns(ws, header_row, name_col)
        for row_idx in range(header_row + 1, ws.max_row + 1):
            record = row_record(ws, lookup, row_idx, headers, name_col, qty_col, spec_col)
            if record:
                level, part_no = row_level_and_part_no(ws, row_idx, hierarchy_cols)
                record["bom_level"] = level
                record["part_no"] = part_no
                record.update({"model": model, "file_name": path.name, "file_path": str(path), "sheet": ws.title, "row_number": row_idx})
                parsed.append(record)
    return model, parsed


def excel_files(folder):
    root = Path(folder)
    if not root.exists():
        return []
    patterns = ["*.xlsx", "*.xlsm", "*.xltx", "*.xltm"]
    files = []
    for pattern in patterns:
        files.extend(root.glob(pattern))
    return [p for p in files if is_real_bom_file(p)]


def is_real_bom_file(path):
    if path.name.startswith("~$"):
        return False
    if normalize_text(path.stem) in HIDDEN_MODEL_NAMES:
        return False
    return path.is_file()


def index_folder(force=False):
    with index_lock:
        if index_state["running"]:
            return
        index_state.update({"running": True, "message": "Scanning BOM folder", "indexed": 0, "total": 0, "errors": []})
    try:
        settings = get_settings()
        files = excel_files(settings["bom_folder"])
        with db() as conn:
            if force:
                conn.execute("DELETE FROM rows")
                conn.execute("DELETE FROM files")
                known = {}
            known = {r["path"]: r for r in conn.execute("SELECT * FROM files")}
            existing = {str(p) for p in files}
            for old_path in set(known) - existing:
                old = known[old_path]
                conn.execute("DELETE FROM rows WHERE file_id=?", (old["id"],))
                conn.execute("DELETE FROM files WHERE path=?", (old_path,))
            to_index = []
            for path in files:
                stat = path.stat()
                current = known.get(str(path))
                changed = not current or current["size"] != stat.st_size or abs(current["mtime"] - stat.st_mtime) > 0.001
                if force or changed:
                    to_index.append(path)
        with index_lock:
            index_state.update({"total": len(to_index), "message": f"Indexing {len(to_index)} changed BOM file(s)"})
        for path in to_index:
            try:
                model, rows = parse_workbook(path)
                stat = path.stat()
                with db() as conn:
                    old = conn.execute("SELECT id FROM files WHERE path=?", (str(path),)).fetchone()
                    if old:
                        conn.execute("DELETE FROM rows WHERE file_id=?", (old["id"],))
                        conn.execute("DELETE FROM files WHERE id=?", (old["id"],))
                    cur = conn.execute(
                        "INSERT INTO files(path,name,model,size,mtime,row_count,indexed_at) VALUES(?,?,?,?,?,?,?)",
                        (str(path), path.name, model, stat.st_size, stat.st_mtime, len(rows), datetime.now().isoformat(timespec="seconds")),
                    )
                    file_id = cur.lastrowid
                    conn.executemany(
                        """
                        INSERT INTO rows(file_id,model,file_name,file_path,sheet,row_number,bom_level,part_no,name_cn,name_vi,pinyin,quantity,specification,headers_json,raw_row_json,search_text,normalized_text)
                        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                        """,
                        [
                            (
                                file_id,
                                r["model"],
                                r["file_name"],
                                r["file_path"],
                                r["sheet"],
                                r["row_number"],
                                r["bom_level"],
                                r["part_no"],
                                r["name_cn"],
                                r["name_vi"],
                                r["pinyin"],
                                r["quantity"],
                                r["specification"],
                                json.dumps(r["headers"], ensure_ascii=False),
                                json.dumps(r["raw"], ensure_ascii=False),
                                r["search_text"],
                                r["normalized_text"],
                            )
                            for r in rows
                        ],
                    )
                with index_lock:
                    index_state["indexed"] += 1
            except Exception as exc:
                with index_lock:
                    index_state["errors"].append(f"{path.name}: {exc}")
        with index_lock:
            index_state["message"] = "Index complete"
    finally:
        with index_lock:
            index_state["running"] = False


def start_index(force=False):
    thread = threading.Thread(target=index_folder, kwargs={"force": force}, daemon=True)
    thread.start()


def expand_terms(query):
    terms = [query]
    normalized_query = normalize_text(query)
    terminology = load_json(TERMINOLOGY_PATH, {})
    exact_matches = [(vi, zh_list) for vi, zh_list in terminology.items() if normalize_text(vi) == normalized_query]
    if exact_matches:
        for vi, zh_list in exact_matches:
            terms.append(vi)
            terms.extend(zh_list)
        seen = []
        for term in terms:
            if term and term not in seen:
                seen.append(term)
        return seen
    for vi, zh_list in terminology.items():
        if normalized_query and phrase_contains(vi, query):
            terms.extend(zh_list)
            terms.append(vi)
        for zh in zh_list:
            if query and query in zh:
                terms.append(vi)
                terms.extend(zh_list)
    seen = []
    for term in terms:
        if term and term not in seen:
            seen.append(term)
    return seen


def score_row(row, query, terms):
    text = row["name_cn"] or ""
    norm_text = normalize_text(text)
    nq = normalize_text(query)
    if query and query in text:
        return 100
    if nq and nq in norm_text:
        return 92
    for term in terms:
        if term and term in text:
            return 88
        nt = normalize_text(term)
        if nt and nt in norm_text:
            return 84
    return SequenceMatcher(None, nq, norm_text).ratio() * 70 if nq else 0


def row_to_dict(row, score=None):
    data = dict(row)
    data["headers"] = json.loads(data.pop("headers_json"))
    data["raw_row"] = json.loads(data.pop("raw_row_json"))
    if score is not None:
        data["score"] = round(score, 1)
    return data


def parse_quantity(value):
    text = str(value or "").strip().replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return Decimal(match.group(0))
    except InvalidOperation:
        return None


def format_quantity(value):
    if value is None:
        return ""
    if value == value.to_integral():
        return str(value.to_integral())
    return format(value.normalize(), "f")


def merge_unique(existing, new_value):
    new_value = str(new_value or "").strip()
    if not new_value:
        return existing
    parts = [part.strip() for part in str(existing or "").split("\n") if part.strip()]
    if new_value not in parts:
        parts.append(new_value)
    return "\n".join(parts)


def aggregate_duplicate_parts(rows):
    grouped = {}
    no_code_index = 0
    for row in rows:
        part_no = str(row.get("part_no") or "").strip()
        if part_no:
            key = (row.get("model"), part_no)
        else:
            no_code_index += 1
            key = (row.get("model"), f"__row_{no_code_index}")
        qty = parse_quantity(row.get("quantity"))
        if key not in grouped:
            merged = dict(row)
            merged["_qty_total"] = qty
            merged["_match_count"] = 1
            grouped[key] = merged
            continue
        merged = grouped[key]
        merged["_match_count"] += 1
        if qty is not None:
            merged["_qty_total"] = (merged.get("_qty_total") or Decimal(0)) + qty
        merged["score"] = max(merged.get("score") or 0, row.get("score") or 0)
        merged["bom_level"] = min(merged.get("bom_level") or row.get("bom_level") or 99, row.get("bom_level") or 99)
        merged["name_cn"] = merge_unique(merged.get("name_cn"), row.get("name_cn"))
        merged["specification"] = merge_unique(merged.get("specification"), row.get("specification"))
        merged["search_text"] = f"{merged.get('search_text', '')} {row.get('search_text', '')}".strip()
    results = []
    for row in grouped.values():
        if row.get("_qty_total") is not None:
            row["quantity"] = format_quantity(row["_qty_total"])
        row.pop("_qty_total", None)
        row.pop("_match_count", None)
        results.append(row)
    return results


def api_models(query, limit=80):
    nq = normalize_text(query)
    if not nq:
        return []
    with db() as conn:
        rows = conn.execute("SELECT id,path,name,model,row_count,indexed_at FROM files ORDER BY name").fetchall()
    scored = []
    for row in rows:
        if normalize_text(row["model"]) in HIDDEN_MODEL_NAMES:
            continue
        hay = normalize_text(row["model"] + " " + row["name"])
        if nq in hay:
            score = 100 if normalize_text(row["model"]) == nq else 90 if hay.startswith(nq) else 75
        else:
            score = SequenceMatcher(None, nq, hay).ratio() * 70
        if score >= 30:
            scored.append((score, row))
    scored.sort(key=lambda item: (-item[0], item[1]["name"]))
    return [dict(r) | {"score": round(s, 1)} for s, r in scored[:limit]]


def visible_file_count(conn):
    rows = conn.execute("SELECT model FROM files").fetchall()
    return sum(1 for row in rows if normalize_text(row["model"]) not in HIDDEN_MODEL_NAMES)


def api_search(payload):
    query = (payload.get("query") or "").strip()
    model_ids = payload.get("model_ids") or ([payload["model_id"]] if payload.get("model_id") else [])
    limit = int(payload.get("limit") or 300)
    terms = expand_terms(query)
    with db() as conn:
        params = []
        where = []
        if model_ids:
            where.append(f"file_id IN ({','.join('?' for _ in model_ids)})")
            params.extend(model_ids)
        sql = "SELECT * FROM rows"
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY file_name,row_number"
        candidates = conn.execute(sql, params).fetchall()
    results = []
    for row in candidates:
        score = score_row(row, query, terms)
        if score >= 42 or (not query and len(results) < limit):
            results.append(row_to_dict(row, score))
    results.sort(key=lambda r: (-r["score"], r["file_name"], r["row_number"]))
    results = aggregate_duplicate_parts(results)
    results.sort(key=lambda r: (-r["score"], r["file_name"], r["row_number"]))
    return {"query": query, "terms": terms, "count": len(results), "results": results[:limit]}


def api_compare(payload):
    response = api_search({"query": payload.get("query", ""), "model_ids": payload.get("model_ids", []), "limit": payload.get("limit", 800)})
    grouped = {}
    for row in response["results"]:
        grouped.setdefault(row["model"], []).append(row)
    keys = {"name_cn": set(), "quantity": set(), "specification": set()}
    for row in response["results"]:
        for key in keys:
            keys[key].add(row.get(key) or "")
    diff_keys = [k for k, vals in keys.items() if len(vals) > 1]
    return response | {"groups": grouped, "diff_keys": diff_keys}


def open_bom(payload):
    path = payload.get("path")
    if not path or not Path(path).exists():
        return {"ok": False, "error": "File not found"}
    row = payload.get("row")
    sheet = payload.get("sheet")
    try:
        if row and sheet:
            ps_path = path.replace("'", "''")
            ps_sheet = str(sheet).replace("'", "''")
            script = (
                "$xl=New-Object -ComObject Excel.Application; "
                "$xl.Visible=$true; "
                f"$wb=$xl.Workbooks.Open('{ps_path}'); "
                f"$ws=$wb.Worksheets.Item('{ps_sheet}'); "
                "$ws.Activate(); "
                f"$ws.Cells.Item({int(row)},1).Select()"
            )
            subprocess.Popen(["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script], close_fds=True)
        else:
            os.startfile(path)
        return {"ok": True}
    except Exception as exc:
        try:
            os.startfile(path)
            return {"ok": True, "warning": str(exc)}
        except Exception as fallback:
            return {"ok": False, "error": str(fallback)}


class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):
        return

    def send_json(self, data, status=200):
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def read_body(self):
        length = int(self.headers.get("Content-Length") or 0)
        if length == 0:
            return {}
        return json.loads(self.rfile.read(length).decode("utf-8"))

    def do_GET(self):
        parsed = urllib.parse.urlparse(self.path)
        query = urllib.parse.parse_qs(parsed.query)
        try:
            if parsed.path == "/api/settings":
                return self.send_json(public_settings())
            if parsed.path == "/api/status":
                with db() as conn:
                    files = visible_file_count(conn)
                    rows = conn.execute("SELECT COUNT(*) c FROM rows").fetchone()["c"]
                return self.send_json({"index": index_state, "files": files, "rows": rows})
            if parsed.path == "/api/models":
                return self.send_json(api_models(query.get("q", [""])[0]))
            if parsed.path == "/api/terminology":
                return self.send_json(load_json(TERMINOLOGY_PATH, {}))
            return self.serve_static(parsed.path)
        except Exception as exc:
            return self.send_json({"error": str(exc)}, 500)

    def do_POST(self):
        try:
            payload = self.read_body()
            if self.path == "/api/settings":
                settings = get_settings()
                if "bom_folder" in payload:
                    settings["bom_folder"] = payload["bom_folder"]
                if "admin_user" in payload:
                    settings["admin_user"] = payload["admin_user"]
                if payload.get("admin_password"):
                    settings["admin_password_hash"] = hashlib.sha256(str(payload["admin_password"]).encode("utf-8")).hexdigest()
                save_json(SETTINGS_PATH, settings)
                return self.send_json(public_settings())
            if self.path == "/api/auth":
                return self.send_json({"ok": check_auth(payload)})
            if self.path == "/api/refresh":
                start_index(False)
                return self.send_json({"ok": True, "message": "Refresh started"})
            if self.path == "/api/reindex":
                start_index(True)
                return self.send_json({"ok": True, "message": "Re-index started"})
            if self.path == "/api/search":
                return self.send_json(api_search(payload))
            if self.path == "/api/compare":
                return self.send_json(api_compare(payload))
            if self.path == "/api/terminology":
                save_json(TERMINOLOGY_PATH, payload)
                return self.send_json(payload)
            if self.path == "/api/open":
                return self.send_json(open_bom(payload))
            return self.send_json({"error": "Not found"}, 404)
        except Exception as exc:
            return self.send_json({"error": str(exc)}, 500)

    def serve_static(self, path):
        safe = "index.html" if path in {"/", ""} else path.lstrip("/")
        target = (STATIC_DIR / safe).resolve()
        if not str(target).startswith(str(STATIC_DIR.resolve())) or not target.exists():
            target = STATIC_DIR / "index.html"
        body = target.read_bytes()
        mime = mimetypes.guess_type(target.name)[0] or "application/octet-stream"
        if target.suffix in {".html", ".css", ".js"}:
            mime += "; charset=utf-8"
        self.send_response(200)
        self.send_header("Content-Type", mime)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)


def main():
    init_db()
    start_index(False)
    port = int(os.environ.get("BOM_TOOL_PORT", "8765"))
    server = ThreadingHTTPServer(("127.0.0.1", port), Handler)
    url = f"http://127.0.0.1:{port}"
    print(f"BOM lookup is running at {url}")
    print("Close this window or press Ctrl+C to stop.")
    if os.environ.get("BOM_TOOL_NO_BROWSER") != "1":
        threading.Timer(0.8, lambda: webbrowser.open(url)).start()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass


if __name__ == "__main__":
    main()
